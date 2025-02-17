import json
from zipfile import BadZipFile

from celery.result import AsyncResult
from django.http import HttpResponseRedirect
from django.urls import reverse_lazy
from django.views.generic import FormView, TemplateView
from openpyxl import load_workbook

from core.celery_app import app

from .forms import LogMeInForm, UploadCreationForm, UploadUpdateForm
from .tasks import check_sirets
from .validator.constants import ETABLISSEMENTS_CREATE_FIELDS, ETABLISSEMENTS_UPDATE_FIELDS, ROLES_FIELDS
from .validator.row_models import EtabCreateRows, EtabUpdateRows, RoleRows


class FileReadingException(Exception):
    pass


class TabException(Exception):
    pass


class InvalidHeaderException(Exception):
    pass


def load_create_xlsx(file):
    try:
        wb = load_workbook(
            filename=file,
            read_only=True,
            keep_links=False,
            data_only=True,
            keep_vba=False,
        )
    except ValueError:
        raise FileReadingException
    sheetnames = wb.sheetnames
    if len(sheetnames) != 2:
        raise TabException
    if sheetnames[0] != "etablissements":
        raise TabException
    if sheetnames[1] != "roles":
        raise TabException
    return wb


def load_update_xlsx(file):
    try:
        wb = load_workbook(
            filename=file,
            read_only=True,
            keep_links=False,
            data_only=True,
            keep_vba=False,
        )
    except ValueError:
        raise FileReadingException
    sheetnames = wb.sheetnames

    if len(sheetnames) != 1:
        raise TabException
    if sheetnames[0] != "etablissements":
        raise TabException

    return wb


def validate_header(first_row, expected_header):
    """Validate first worksheet row matches `expected_header`"""

    header = [cell.value for cell in first_row]

    if header != expected_header:
        raise InvalidHeaderException


class ValidateCreationFileView(FormView):
    """
    Performs form submission and main validation.
    If validation fails, display error messages in a table
    If validation succeeds, launch an async task and redirects to a view polling results from the queue.
    """

    form_class = UploadCreationForm
    template_name = "mass_validator/validate_create.html"

    @property
    def has_errors(self):
        return self.errors or self.parse_error or self.enough_rows_error or self.too_many_rows_error

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.errors = []
        self.parse_error = False
        self.parse_error_message = ""
        self.enough_rows_error = False
        self.too_many_rows_error = False
        self.async_task_id = None

    def check_sirets_exists(self, etab_rows):
        to_check = [{"siret": row.siret, "row_number": row.index} for row in etab_rows]

        async_task = check_sirets.delay(to_check)

        self.async_task_id = async_task.id

    def parse(self, file):
        try:
            wb = load_create_xlsx(file)
        except (BadZipFile, KeyError, TabException, FileReadingException):
            self.parse_error = True
            return

        ws_etablissements = wb.worksheets[0]

        ws_roles = wb.worksheets[1]
        etab_first_row = ws_etablissements[1][: len(ETABLISSEMENTS_CREATE_FIELDS)]
        role_first_row = ws_roles[1][: len(ROLES_FIELDS)]

        # performs header validation, exits if it fails
        try:
            validate_header(etab_first_row, ETABLISSEMENTS_CREATE_FIELDS)
            validate_header(role_first_row, ROLES_FIELDS)
        except InvalidHeaderException:
            self.parse_error = True
            return

        etab_rows = EtabCreateRows.from_worksheet(ws_etablissements)

        etab_rows.validate()

        # exits if customer is too lazy
        if not etab_rows.has_enough_rows:
            self.enough_rows_error = True
            return

        # exits if too many rows
        if etab_rows.has_too_many_rows:
            self.too_many_rows_error = True
            return

        role_rows = RoleRows.from_worksheet(ws_roles)

        role_rows.validate(etab_rows.sirets())

        # main validation
        if not etab_rows.is_valid:
            self.errors.extend(etab_rows.get_errors())
        if not role_rows.is_valid:
            self.errors.extend(role_rows.get_errors())

        # This validation can occur when both tabs are already validated
        if etab_rows.is_valid and role_rows.is_valid:
            etab_rows.validate_have_admin(role_rows.admin_sirets())
            if not etab_rows.is_valid:
                self.errors.extend(etab_rows.get_errors())

        # only performs api checks if everything else passes
        if not self.has_errors:
            # async task
            self.check_sirets_exists(etab_rows)

    def form_valid(self, form):
        file = self.request.FILES["file"]

        self.parse(file)

        if self.has_errors:
            return self.error_page()
        # if form valid, redirect to a page polling siret api validation
        return HttpResponseRedirect(self.get_success_url())

    def error_page(self):
        return self.render_to_response(
            {
                "errors": self.errors,
                "parse_error": self.parse_error,
                "has_errors": self.has_errors,
                "enough_rows_error": self.enough_rows_error,
                "too_many_rows_error": self.too_many_rows_error,
            }
        )

    def get_success_url(self):
        if self.async_task_id:
            return reverse_lazy("pollable_result", args=[self.async_task_id])
        return reverse_lazy("create_result")


class CreateResultView(TemplateView):
    """Optional `task_id` trigger result polling in template"""

    template_name = "mass_validator/create_result.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update({"task_id": self.kwargs.get("task_id", None)})
        return ctx


STATE_RUNNING = "running"
STATE_DONE = "done"


class CheckSiretView(TemplateView):
    """View to be called by CreateResultView template to render api call results when done"""

    template_name = "mass_validator/_sirets_result.html"

    def dispatch(self, request, *args, **kwargs):
        self.task_id = self.kwargs.get("task_id")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        job = AsyncResult(self.task_id, app=app)
        done = job.ready()

        result = job.result

        if isinstance(result, dict):
            progress = result.get("progress", 0)
        else:
            progress = 100.0 if done else 0.0

        ctx.update({"progress": progress})

        if not job.ready():
            ctx.update({"state": STATE_RUNNING})
        else:
            ctx.update({"siret_errors": job.get(), "state": STATE_DONE})
        return ctx


class ValidateUpdateFileView(FormView):
    form_class = UploadUpdateForm
    template_name = "mass_validator/validate_update.html"
    success_url = "/"

    @property
    def has_errors(self):
        return self.errors or self.parse_error or self.enough_rows_error or self.too_many_rows_error

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.errors = []
        self.parse_error = False
        self.parse_error_message = ""
        self.enough_rows_error = False
        self.too_many_rows_error = False
        self.async_task_id = None
        self.json_export = []

    def error_page(self):
        return self.render_to_response(
            {
                "errors": self.errors,
                "parse_error": self.parse_error,
                "has_errors": self.has_errors,
                "enough_rows_error": self.enough_rows_error,
                "too_many_rows_error": self.too_many_rows_error,
            }
        )

    def success_page(self):
        kwargs = {"success": True}
        if self.request.connected:
            kwargs["json_export"] = json.dumps(self.json_export, indent=4)

        return self.render_to_response(kwargs)

    def parse(self, file):
        try:
            wb = load_update_xlsx(file)
        except (BadZipFile, KeyError, TabException, FileReadingException):
            self.parse_error = True
            return

        ws_etablissements = wb.worksheets[0]

        etab_first_row = ws_etablissements[1][: len(ETABLISSEMENTS_UPDATE_FIELDS)]

        # performs header validation, exits if it fails
        try:
            validate_header(etab_first_row, ETABLISSEMENTS_UPDATE_FIELDS)

        except InvalidHeaderException:
            self.parse_error = True
            return

        etab_rows = EtabUpdateRows.from_worksheet(ws_etablissements)

        etab_rows.validate()

        # exits if customer is too lazy
        if not etab_rows.has_enough_rows:
            self.enough_rows_error = True
            return

        # exits if too many rows
        if etab_rows.has_too_many_rows:
            self.too_many_rows_error = True
            return

        if not etab_rows.is_valid:
            self.errors.extend(etab_rows.get_errors())
            return

        self.json_export = etab_rows.as_json()

    def form_valid(self, form):
        file = self.request.FILES["file"]

        self.parse(file)

        if self.has_errors:
            return self.error_page()

        return self.success_page()


class UpdateResultView(TemplateView):
    template_name = "mass_validator/update_result.html"


cookie_max_age = 60 * 24 * 30  # one month


class LogMeIn(FormView):
    form_class = LogMeInForm
    template_name = "mass_validator/log_me_in.html"
    success_url = "/"

    def form_valid(self, form):
        res = super().form_valid(form)
        res.set_signed_cookie("validator_connected", "connected", max_age=cookie_max_age)
        return res
